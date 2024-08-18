'''
Script to generate cuts report in excel file , accepts date_from, date_to returns excel ,file path
'''
import pandas as pd
import mysql.connector as conn
from mysql.connector import Error
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os

connection = conn.connect(
    host="localhost",
    user="backend_comm",
    database="bkt_prms",
    port ="3306",
    password ="ASD-90K-10B8"
    )

date_from = "2024-08-06"
date_to = "2024-08-08"

def downloadReport (date_from:str, date_to:str, shift:str):
    date_from = date_from + " 00:00:00"
    if date_to != date_from:
        date_to += " 23:59:59"
    else:
        date_to += " 00:00:00"
    
    if connection.is_connected():
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM `recipe_report` WHERE `DateTime`> %s  AND `DateTime` < %s", [date_from,date_to])
        result= cursor.fetchall()
        print(type(result))
        column_names = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(result, columns=column_names)
        axc = tuple(column_names)
        cursor.close()
        connection.close()
        
        path = r"BKT_project\PRMS_Reports\report_"+ datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
        write_data_to_xlsx(data=df) 
        df.to_excel(path, index=False)
        
        # output = r"BKT_project\PRMS_Reports\pivot_"+ datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
        # create_pivot_table(input_file_path=path,output_file_path=output )
        return path


def write_data_to_xlsx( data):
        """
        Opens an XLSX file, writes data starting from cell A2,
        saves it at the same location with a different name,
        without changing its format.
        Parameters:
        original_file_path (str): Path to the original XLSX file.
        new_file_name (str): Name of the new XLSX file.
        data (pd.DataFrame): Data to be written to the XLSX file.
        """
        # Load the original workbook
        wb = load_workbook(filename="BKT_project\\PRMS_Reports\\test.xlsx")
        
        # Select the active worksheet
        ws = wb['Sheet1']
        
        # Clear any existing content below A1
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column):
            for cell in row:
                cell.value = None
                
        # Write the new data starting from cell A2
        
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Save the workbook under a new name
        wb.save(filename="BKT_project\\PRMS_Reports\\test2.xlsx")

print(downloadReport(date_from=date_from, date_to=date_to,shift='a'))


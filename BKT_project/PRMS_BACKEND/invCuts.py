import mysql.connector as conn
from openpyxl import load_workbook


def read_excel_conv(filepath:str):
    wb = load_workbook(filename=filepath)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value for cell in row])
    return data
    
def uploadInvCuts(path:str):
    connection = conn.connect(
    host="localhost",
    user="backend_comm",
    database="bkt_prms",
    port ="3306",
    password ="ASD-90K-10B8"
    )
    
    inv_data = read_excel_conv(filepath=path)
    if connection.is_connected():
        print("Connected to MySQL database")
        # Create a cursor object to execute SQL queries
        cursor = connection.cursor()
        cursor.execute("DELETE FROM `recipe_inv` WHERE 1")
        inv_data = inv_data[1:]
        for row in inv_data:
            values = tuple(row)
            cursor.execute("INSERT INTO `recipe_inv` VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",values)
            print(f'inv cuts inserted for id:{values[0]}')
        return True
    
# uploadInvCuts(path='inv_cuts\\inv_1.xlsx')
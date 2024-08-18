import mysql.connector as conn
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from datetime import datetime
from PRMS_BACKEND.setCuts import setCutsUpdate
# FILE_PATH ="D:\\BKT_PRMS\\RECIPE_1.xlsx"
# Connect to the MySQL database

def read_excel_data(filepath:str):
    # Load the workbook
    workbook = load_workbook(filename=filepath)
    # Select the first sheet
    sheet1 = workbook['select_data']
    # Read all data from the sheet
    select_data = []
    for row in sheet1.iter_rows(values_only=True):
        select_data.append(row)
    conv_data = []
    sheet2 = workbook['conv_data']
    for row in sheet2.iter_rows():
        conv_data.append([cell.value for cell in row])

    # Evaluate formulas in the second sheet
    for row in conv_data:
        for i, cell_value in enumerate(row):
            if isinstance(cell_value, str) and cell_value.startswith('='):
                # Evaluate the formula
                formula = cell_value
                result = Tokenizer(formula).evaluate()
                row[i] = result
    width_data = []
    sheet3 = workbook['width_data']
    for row in sheet3.iter_rows(values_only=True):
        width_data.append(row)

    return select_data, conv_data, width_data
# Check if the connection was successful
# script to update and insert schedule on server

def updateRecipe(FILE_PATH:str):
    connection = conn.connect(
    host="localhost",
    user="backend_comm",
    database="bkt_prms",
    port ="3306",
    password ="ASD-90K-10B8"
)
    if connection.is_connected():
        print("Connected to MySQL database")
        # Create a cursor object to execute SQL queries
        cursor = connection.cursor()
        selectData, convData, widthData = read_excel_data(FILE_PATH)
        data = selectData [1:]   
        for row in data :
            prod_id = row[1] # prod_id is the 2nd column
            values = tuple(row[2:])   # drop SR_NO as its a primary key
            cursor.execute("SELECT COUNT(*) FROM recipe_1_select WHERE PROD_ID = %s", (prod_id,))
            result = cursor.fetchone()[0]
            if result>0:
                try: 
                    cursor.execute("""
                                UPDATE recipe_1_select
                                SET SIZE = %s, FABRIC_PLY_1 = %s, FABRIC_PLY_2 = %s, FABRIC_PLY_3 = %s, FABRIC_PLY_4 = %s, FABRIC_PLY_5 = %s, FABRIC_PLY_6 = %s, FABRIC_PLY_7 = %s, FABRIC_PLY_8 = %s, FABRIC_PLY_9 = %s, FABRIC_PLY_10 = %s, FABRIC_BRAKER_11 = %s, FABRIC_BRAKER_12 = %s, FABRIC_BRAKER_13 = %s, FABRIC_BRAKER_14 = %s, NO_OF_TYRES = %s WHERE PROD_ID = %s """, 
                    values + (prod_id,))
                except Exception as e:
                    print(f"error in update query SR_NO:{row[0]}, PROD_ID:{prod_id}, \n Error-->{e}")
            else:
                try:
                    # PROD_ID does not exist, insert a new row
                    cursor.execute("""
                INSERT INTO recipe_1_select (PROD_ID, SIZE, FABRIC_PLY_1, FABRIC_PLY_2, FABRIC_PLY_3, FABRIC_PLY_4, FABRIC_PLY_5, FABRIC_PLY_6, FABRIC_PLY_7, FABRIC_PLY_8, FABRIC_PLY_9,FABRIC_PLY_10, FABRIC_BRAKER_11,FABRIC_BRAKER_12, FABRIC_BRAKER_13, FABRIC_BRAKER_14,NO_OF_TYRES) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,  (prod_id,)+values )
                except Exception as e:
                    print( f"Insertion failed at recipe_1_select for SR_NO :{row[0]}, PROD_ID: {row[1]}, error--> {e}")

        data = convData[1:]
        for row in data:
            prod_id = row[0]
            values = tuple(row[1:]) 
            cursor.execute("SELECT COUNT(*) FROM recipe_conv_table WHERE PROD_ID = %s", (prod_id,))
            result = cursor.fetchone()[0]
            if result>0:
                try:
                    cursor.execute("""
                                UPDATE recipe_conv_table SET
                                Angle_1 = %s, Angle_2= %s, Angle_3 = %s, Angle_4 = %s, Angle_5 = %s, Angle_6 = %s, Angle_7 = %s, Angle_8=%s, Angle_9 =%s, Angle_10=%s, Angle_11 = %s, Angle_12 = %s, Angle_13= %s, Angle_14 = %s, CONV_1 = %s, CONV_2 = %s, CONV_3 = %s, CONV_4 = %s, CONV_5 = %s,CONV_6 = %s, CONV_7 = %s, CONV_8 = %s, CONV_9 = %s, CONV_10 = %s, CONV_11 = %s , CONV_12 = %s, CONV_13 = %s, CONV_14 = %s WHERE PROD_ID = %s
                                """, values+(prod_id,))
                    print(f"updated in conv_table at PROD_ID:{prod_id}")
                except Exception as e:
                    print(f"Error in updating conv table prod_id: {prod_id}, \n error-->{e}")
            else:
                try:
                    # print(f"inserting conv current row id:{prod_id}")
                    if prod_id:
                        cursor.execute("""INSERT INTO recipe_conv_table 
                            (PROD_ID,Angle_1, Angle_2, Angle_3, Angle_4, Angle_5, Angle_6, Angle_7, Angle_8, Angle_9, Angle_10, Angle_11,Angle_12, Angle_13, Angle_14, CONV_1, CONV_2, CONV_3 ,CONV_4,CONV_5, CONV_6, CONV_7, CONV_8, CONV_9, CONV_10, CONV_11, CONV_12, CONV_13, CONV_14 ) VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s, %s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s)""", (prod_id,)+values)
                        print(f"Inserted  in conv_table at PROD_ID:{prod_id}")
                except Exception as e:
                    print(f"error in insterting conv table at prod_id:{prod_id}, \n error-->{e}")
                    
        data = widthData[1:]
        temp_list = []
        for row in data:
            prod_id = row[0]
            values = tuple(row[1:])    
            cursor.execute("SELECT COUNT(*) FROM recipe_width WHERE PROD_ID = %s", (prod_id,))
            result = cursor.fetchone()[0]
            if result>0:
                try: 
                    cursor.execute("""
                                UPDATE recipe_width SET
                                WIDTH_1 = %s, WIDTH_2= %s, WIDTH_3 = %s, WIDTH_4 = %s, WIDTH_5 = %s, WIDTH_6 = %s, WIDTH_7 = %s, WIDTH_8=%s, WIDTH_9 =%s, WIDTH_10=%s, WIDTH_11 = %s, WIDTH_12 = %s, WIDTH_13= %s, WIDTH_14 = %s WHERE PROD_ID = %s
                                """, values+(prod_id,))
                    print(f"updated in recipe_width at PROD_ID:{prod_id}")
                    temp_list.append(str(prod_id))
                except Exception as e:
                    print(f"Error in updating recipe_width prod_id: {prod_id}, \n error-->{e}")
            else:
                try:
                    cursor.execute("""INSERT INTO recipe_width
                            (PROD_ID,WIDTH_1, WIDTH_2, WIDTH_3, WIDTH_4, WIDTH_5, WIDTH_6, WIDTH_7, WIDTH_8, WIDTH_9, WIDTH_10, WIDTH_11,WIDTH_12, WIDTH_13, WIDTH_14 ) VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s,%s, %s)""", (prod_id,)+values)
                    print(f"Inserted in recipe_width at PROD_ID:{prod_id}")
                except Exception as e:
                    print(f"error :Insterting in recipe_width table at prod_id:{prod_id}, \n error-->{e}")

        # Update the set_cuts , set_Tyres  as per selectData
        setCutsUpdate(connection=connection,selectData=selectData)
        # Close the cursor and connection
        cursor.close()
        connection.close()
        print("MySQL connection is closed")
        return 1
    else:      
        print("Failed to connect to MySQL database")
        return 0

    # delete the reecipe xlsx



def saveLogs(log:str):
    try:
        # Load the workbook
        wb = load_workbook("C:\\BKT_Bias_cutter_3\\Book1.xlsx")
        # Select the worksheet
        ws = wb["Sheet1"]
        # Create a timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([timestamp,log])
        wb.save("C:\\BKT_Bias_cutter_3\\Book1.xlsx")
        return 1  # Return 1 to indicate success
    except Exception as e:
        print(f"An error occurred while saving logs: {e}")
        return 0  # Return 0 to indicate failure

'''
Script to generate cuts report in excel file , accepts date_from, date_to returns excel ,file path
'''
import mysql.connector as conn
from mysql.connector import Error
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import math

SAVE_DIR = "C:\\Reports"

def downloadCutReport(date_from:str, date_to:str, shift='A'):
    date_from = date_from + " 00:00:00"
    if date_to != date_from:
        date_to += " 23:59:59"
    else:
        date_to += " 00:00:00"

    connection = conn.connect(
    host="localhost",
    user="backend_comm",
    database="bkt_prms",
    port ="3306",
    password ="ASD-90K-10B8"
    )
    try:
        if connection.is_connected():
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM `recipe_report` WHERE `DateTime`> %s  AND `DateTime` < %s", [date_from,date_to])
            result= cursor.fetchall()
            print(type(result))
            column_names = [desc[0] for desc in cursor.description]
            cursor.close()
            connection.close()
            wb = load_workbook(filename='BKT_project/PRMS_Reports/cutReport_format.xlsx')
            ws = wb['Sheet1']

            # Retrieve the table
            table = ws.tables['Table1']

            # Get the starting row of the table's data (first row after column names)
            start_row = int(table.ref.split(':')[0][1:]) + 1  # Data starts after header
            start_col_letter = table.ref.split(':')[0][0]

            # Find the first empty row in the table
            first_empty_row = start_row
            while ws.cell(row=first_empty_row, column=ws[start_col_letter + str(first_empty_row)].column).value is not None:
                first_empty_row += 1

            # Write data inside the table starting from the first empty row using ws.append()
            for row_tuple in result:
                row = list(row_tuple)
                row[-2] = math.ceil(row[-2]) # round off remaining tyres
                row[-3] = math.ceil(row[-3]) # round off act tyres
                ws.append(row)

            # Function to check if a row is empty
            def is_row_empty(ws, row_num):
                return all(ws.cell(row=row_num, column=col).value is None for col in range(1, ws.max_column + 1))
            
            # Delete empty rows in the range of interest
            for row_num in range(start_row, ws.max_row + 1):
                if is_row_empty(ws, row_num):
                    ws.delete_rows(row_num)
            
            # Update the table's reference to include the new rows
            new_end_row = first_empty_row + len(result) -1
            end_col_letter = table.ref.split(':')[1][0]  # Get the last column letter
            table.ref = f"{start_col_letter}{start_row-1}:{end_col_letter}{new_end_row}"
            ws = wb['Sheet2']
            ws['H1'] = datetime.now().strftime("Date: %Y-%m-%d   Time:%H:%M:%S") 
            ws['H1'].font = Font(name='Cascadia Code SemiBold', size=14)
            # Save the workbook
            path ="C:\\BiasCutterReport\\report_new_"+ datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"
            wb.save(filename=path)
            
            print(f"report saved as : path[25:]")    
            return (path)  
    except Exception as e:
        print(f"error occured as:{e}")  
        return -1  